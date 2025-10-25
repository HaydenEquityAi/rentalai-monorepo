"""
Financial Report Export Service
Professional PDF and Excel generation for financial reports
"""
from io import BytesIO
from datetime import date
from decimal import Decimal
from typing import Dict, Any
import logging

# PDF Generation
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT

# Excel Generation
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins

logger = logging.getLogger(__name__)


class ReportExportService:
    """Service for exporting financial reports to PDF and Excel formats"""

    @staticmethod
    def _format_currency(amount: str) -> str:
        """Format currency as $X,XXX.XX"""
        try:
            decimal_amount = Decimal(amount)
            return f"${decimal_amount:,.2f}"
        except (ValueError, TypeError):
            return "$0.00"

    @staticmethod
    def _get_pdf_styles():
        """Get professional PDF styles"""
        styles = getSampleStyleSheet()
        
        # Title style
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=12,
            alignment=TA_CENTER,
            textColor=colors.darkblue
        )
        
        # Subtitle style
        subtitle_style = ParagraphStyle(
            'CustomSubtitle',
            parent=styles['Heading2'],
            fontSize=12,
            spaceAfter=20,
            alignment=TA_CENTER,
            textColor=colors.darkgrey
        )
        
        # Section header style
        section_style = ParagraphStyle(
            'CustomSection',
            parent=styles['Heading3'],
            fontSize=14,
            spaceAfter=8,
            spaceBefore=12,
            textColor=colors.darkblue
        )
        
        return title_style, subtitle_style, section_style

    @staticmethod
    def _create_pdf_table(data: list, col_widths: list = None) -> Table:
        """Create a professionally styled PDF table"""
        if col_widths is None:
            col_widths = [3*inch, 1.5*inch]
        
        table = Table(data, colWidths=col_widths)
        
        # Professional table styling
        table_style = TableStyle([
            # Header styling
            ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            
            # Data styling
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('ALIGN', (0, 1), (0, -1), 'LEFT'),  # Left align descriptions
            ('ALIGN', (1, 1), (-1, -1), 'RIGHT'),  # Right align amounts
            
            # Grid lines
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            
            # Alternating row colors
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
            
            # Total row styling
            ('BACKGROUND', (0, -1), (-1, -1), colors.lightblue),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, -1), (-1, -1), 11),
        ])
        
        table.setStyle(table_style)
        return table

    @staticmethod
    def _get_excel_styles():
        """Get professional Excel styles"""
        # Header style
        header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        # Data style
        data_font = Font(name='Arial', size=10)
        data_alignment = Alignment(horizontal='right', vertical='center')
        
        # Currency style
        currency_font = Font(name='Arial', size=10)
        currency_alignment = Alignment(horizontal='right', vertical='center')
        
        # Total style
        total_font = Font(name='Arial', size=10, bold=True)
        total_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        
        # Border
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        return {
            'header_font': header_font,
            'header_fill': header_fill,
            'header_alignment': header_alignment,
            'data_font': data_font,
            'data_alignment': data_alignment,
            'currency_font': currency_font,
            'currency_alignment': currency_alignment,
            'total_font': total_font,
            'total_fill': total_fill,
            'border': thin_border
        }

    @staticmethod
    def export_profit_loss_pdf(data: Dict[str, Any], org_name: str, start_date: date, end_date: date) -> BytesIO:
        """Export Profit & Loss statement to PDF"""
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=1*inch)
        story = []
        
        title_style, subtitle_style, section_style = ReportExportService._get_pdf_styles()
        
        # Title and subtitle
        story.append(Paragraph("PROFIT & LOSS STATEMENT", title_style))
        story.append(Paragraph(f"{org_name}", subtitle_style))
        story.append(Paragraph(f"For the period {start_date.strftime('%B %d, %Y')} to {end_date.strftime('%B %d, %Y')}", subtitle_style))
        story.append(Spacer(1, 20))
        
        # Revenue section
        story.append(Paragraph("REVENUE", section_style))
        revenue_data = [["Description", "Amount"]]
        revenue_data.append(["Total Revenue", ReportExportService._format_currency(data.get("revenue", "0"))])
        revenue_table = ReportExportService._create_pdf_table(revenue_data)
        story.append(revenue_table)
        story.append(Spacer(1, 12))
        
        # Expenses section
        story.append(Paragraph("EXPENSES", section_style))
        expenses_data = [["Description", "Amount"]]
        expenses_data.append(["Total Expenses", ReportExportService._format_currency(data.get("expenses", "0"))])
        expenses_table = ReportExportService._create_pdf_table(expenses_data)
        story.append(expenses_table)
        story.append(Spacer(1, 12))
        
        # Net Income section
        story.append(Paragraph("NET INCOME", section_style))
        net_income_data = [["Description", "Amount"]]
        net_income_data.append(["Net Income", ReportExportService._format_currency(data.get("net_income", "0"))])
        
        # Special styling for net income
        net_income_table = Table(net_income_data, colWidths=[3*inch, 1.5*inch])
        net_income_style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 1), (-1, 1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 1), (-1, 1), 12),
            ('ALIGN', (0, 1), (0, 1), 'LEFT'),
            ('ALIGN', (1, 1), (1, 1), 'RIGHT'),
            ('BACKGROUND', (0, 1), (-1, 1), colors.lightgreen),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ])
        net_income_table.setStyle(net_income_style)
        story.append(net_income_table)
        
        # Gross profit margin
        if data.get("gross_profit_margin"):
            story.append(Spacer(1, 12))
            margin_data = [["Gross Profit Margin", f"{data['gross_profit_margin']}%"]]
            margin_table = ReportExportService._create_pdf_table(margin_data)
            story.append(margin_table)
        
        doc.build(story)
        buffer.seek(0)
        return buffer

    @staticmethod
    def export_profit_loss_excel(data: Dict[str, Any], org_name: str, start_date: date, end_date: date) -> BytesIO:
        """Export Profit & Loss statement to Excel"""
        buffer = BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = "Profit & Loss"
        
        styles = ReportExportService._get_excel_styles()
        
        # Set page margins
        ws.page_margins = PageMargins(left=0.7, right=0.7, top=0.75, bottom=0.75)
        
        # Title
        ws['A1'] = "PROFIT & LOSS STATEMENT"
        ws['A1'].font = Font(name='Arial', size=16, bold=True, color='366092')
        ws['A1'].alignment = Alignment(horizontal='center')
        ws.merge_cells('A1:B1')
        
        # Organization and period
        ws['A2'] = org_name
        ws['A2'].font = Font(name='Arial', size=12, bold=True)
        ws['A2'].alignment = Alignment(horizontal='center')
        ws.merge_cells('A2:B2')
        
        ws['A3'] = f"For the period {start_date.strftime('%B %d, %Y')} to {end_date.strftime('%B %d, %Y')}"
        ws['A3'].font = Font(name='Arial', size=10)
        ws['A3'].alignment = Alignment(horizontal='center')
        ws.merge_cells('A3:B3')
        
        # Revenue section
        row = 5
        ws[f'A{row}'] = "REVENUE"
        ws[f'A{row}'].font = Font(name='Arial', size=12, bold=True, color='366092')
        row += 1
        
        ws[f'A{row}'] = "Total Revenue"
        ws[f'B{row}'] = float(data.get("revenue", "0"))
        ws[f'B{row}'].number_format = '"$"#,##0.00'
        ws[f'A{row}'].font = styles['data_font']
        ws[f'B{row}'].font = styles['currency_font']
        ws[f'A{row}'].alignment = styles['data_alignment']
        ws[f'B{row}'].alignment = styles['currency_alignment']
        row += 2
        
        # Expenses section
        ws[f'A{row}'] = "EXPENSES"
        ws[f'A{row}'].font = Font(name='Arial', size=12, bold=True, color='366092')
        row += 1
        
        ws[f'A{row}'] = "Total Expenses"
        ws[f'B{row}'] = float(data.get("expenses", "0"))
        ws[f'B{row}'].number_format = '"$"#,##0.00'
        ws[f'A{row}'].font = styles['data_font']
        ws[f'B{row}'].font = styles['currency_font']
        ws[f'A{row}'].alignment = styles['data_alignment']
        ws[f'B{row}'].alignment = styles['currency_alignment']
        row += 2
        
        # Net Income section
        ws[f'A{row}'] = "NET INCOME"
        ws[f'A{row}'].font = Font(name='Arial', size=12, bold=True, color='366092')
        row += 1
        
        ws[f'A{row}'] = "Net Income"
        ws[f'B{row}'] = float(data.get("net_income", "0"))
        ws[f'B{row}'].number_format = '"$"#,##0.00'
        ws[f'A{row}'].font = styles['total_font']
        ws[f'B{row}'].font = styles['total_font']
        ws[f'A{row}'].fill = styles['total_fill']
        ws[f'B{row}'].fill = styles['total_fill']
        ws[f'A{row}'].alignment = styles['data_alignment']
        ws[f'B{row}'].alignment = styles['currency_alignment']
        row += 2
        
        # Gross profit margin
        if data.get("gross_profit_margin"):
            ws[f'A{row}'] = "Gross Profit Margin"
            ws[f'B{row}'] = f"{data['gross_profit_margin']}%"
            ws[f'A{row}'].font = styles['data_font']
            ws[f'B{row}'].font = styles['data_font']
            ws[f'A{row}'].alignment = styles['data_alignment']
            ws[f'B{row}'].alignment = styles['currency_alignment']
        
        # Apply borders
        for row_num in range(1, row + 1):
            for col in ['A', 'B']:
                cell = ws[f'{col}{row_num}']
                cell.border = styles['border']
        
        # Auto-fit columns
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        
        # Freeze panes
        ws.freeze_panes = 'A5'
        
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    @staticmethod
    def export_balance_sheet_pdf(data: Dict[str, Any], org_name: str, as_of_date: date) -> BytesIO:
        """Export Balance Sheet to PDF"""
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=1*inch)
        story = []
        
        title_style, subtitle_style, section_style = ReportExportService._get_pdf_styles()
        
        # Title and subtitle
        story.append(Paragraph("BALANCE SHEET", title_style))
        story.append(Paragraph(f"{org_name}", subtitle_style))
        story.append(Paragraph(f"As of {as_of_date.strftime('%B %d, %Y')}", subtitle_style))
        story.append(Spacer(1, 20))
        
        # Assets section
        story.append(Paragraph("ASSETS", section_style))
        assets_data = [["Description", "Amount"]]
        assets_data.append(["Total Assets", ReportExportService._format_currency(data.get("assets", "0"))])
        assets_table = ReportExportService._create_pdf_table(assets_data)
        story.append(assets_table)
        story.append(Spacer(1, 12))
        
        # Liabilities section
        story.append(Paragraph("LIABILITIES", section_style))
        liabilities_data = [["Description", "Amount"]]
        liabilities_data.append(["Total Liabilities", ReportExportService._format_currency(data.get("liabilities", "0"))])
        liabilities_table = ReportExportService._create_pdf_table(liabilities_data)
        story.append(liabilities_table)
        story.append(Spacer(1, 12))
        
        # Equity section
        story.append(Paragraph("EQUITY", section_style))
        equity_data = [["Description", "Amount"]]
        equity_data.append(["Total Equity", ReportExportService._format_currency(data.get("equity", "0"))])
        equity_table = ReportExportService._create_pdf_table(equity_data)
        story.append(equity_table)
        story.append(Spacer(1, 12))
        
        # Balance verification
        story.append(Paragraph("BALANCE VERIFICATION", section_style))
        balance_data = [["Description", "Amount"]]
        balance_data.append(["Total Liabilities + Equity", ReportExportService._format_currency(data.get("total_liabilities_and_equity", "0"))])
        balance_data.append(["Balance Check (Assets - Liabilities - Equity)", ReportExportService._format_currency(data.get("balance_check", "0"))])
        
        balance_table = Table(balance_data, colWidths=[3*inch, 1.5*inch])
        balance_style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('ALIGN', (0, 1), (0, -1), 'LEFT'),
            ('ALIGN', (1, 1), (-1, -1), 'RIGHT'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
        ])
        balance_table.setStyle(balance_style)
        story.append(balance_table)
        
        doc.build(story)
        buffer.seek(0)
        return buffer

    @staticmethod
    def export_balance_sheet_excel(data: Dict[str, Any], org_name: str, as_of_date: date) -> BytesIO:
        """Export Balance Sheet to Excel"""
        buffer = BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = "Balance Sheet"
        
        styles = ReportExportService._get_excel_styles()
        
        # Set page margins
        ws.page_margins = PageMargins(left=0.7, right=0.7, top=0.75, bottom=0.75)
        
        # Title
        ws['A1'] = "BALANCE SHEET"
        ws['A1'].font = Font(name='Arial', size=16, bold=True, color='366092')
        ws['A1'].alignment = Alignment(horizontal='center')
        ws.merge_cells('A1:B1')
        
        # Organization and date
        ws['A2'] = org_name
        ws['A2'].font = Font(name='Arial', size=12, bold=True)
        ws['A2'].alignment = Alignment(horizontal='center')
        ws.merge_cells('A2:B2')
        
        ws['A3'] = f"As of {as_of_date.strftime('%B %d, %Y')}"
        ws['A3'].font = Font(name='Arial', size=10)
        ws['A3'].alignment = Alignment(horizontal='center')
        ws.merge_cells('A3:B3')
        
        # Assets section
        row = 5
        ws[f'A{row}'] = "ASSETS"
        ws[f'A{row}'].font = Font(name='Arial', size=12, bold=True, color='366092')
        row += 1
        
        ws[f'A{row}'] = "Total Assets"
        ws[f'B{row}'] = float(data.get("assets", "0"))
        ws[f'B{row}'].number_format = '"$"#,##0.00'
        ws[f'A{row}'].font = styles['data_font']
        ws[f'B{row}'].font = styles['currency_font']
        ws[f'A{row}'].alignment = styles['data_alignment']
        ws[f'B{row}'].alignment = styles['currency_alignment']
        row += 2
        
        # Liabilities section
        ws[f'A{row}'] = "LIABILITIES"
        ws[f'A{row}'].font = Font(name='Arial', size=12, bold=True, color='366092')
        row += 1
        
        ws[f'A{row}'] = "Total Liabilities"
        ws[f'B{row}'] = float(data.get("liabilities", "0"))
        ws[f'B{row}'].number_format = '"$"#,##0.00'
        ws[f'A{row}'].font = styles['data_font']
        ws[f'B{row}'].font = styles['currency_font']
        ws[f'A{row}'].alignment = styles['data_alignment']
        ws[f'B{row}'].alignment = styles['currency_alignment']
        row += 2
        
        # Equity section
        ws[f'A{row}'] = "EQUITY"
        ws[f'A{row}'].font = Font(name='Arial', size=12, bold=True, color='366092')
        row += 1
        
        ws[f'A{row}'] = "Total Equity"
        ws[f'B{row}'] = float(data.get("equity", "0"))
        ws[f'B{row}'].number_format = '"$"#,##0.00'
        ws[f'A{row}'].font = styles['data_font']
        ws[f'B{row}'].font = styles['currency_font']
        ws[f'A{row}'].alignment = styles['data_alignment']
        ws[f'B{row}'].alignment = styles['currency_alignment']
        row += 2
        
        # Balance verification
        ws[f'A{row}'] = "BALANCE VERIFICATION"
        ws[f'A{row}'].font = Font(name='Arial', size=12, bold=True, color='366092')
        row += 1
        
        ws[f'A{row}'] = "Total Liabilities + Equity"
        ws[f'B{row}'] = float(data.get("total_liabilities_and_equity", "0"))
        ws[f'B{row}'].number_format = '"$"#,##0.00'
        ws[f'A{row}'].font = styles['data_font']
        ws[f'B{row}'].font = styles['currency_font']
        ws[f'A{row}'].alignment = styles['data_alignment']
        ws[f'B{row}'].alignment = styles['currency_alignment']
        row += 1
        
        ws[f'A{row}'] = "Balance Check (Assets - Liabilities - Equity)"
        ws[f'B{row}'] = float(data.get("balance_check", "0"))
        ws[f'B{row}'].number_format = '"$"#,##0.00'
        ws[f'A{row}'].font = styles['data_font']
        ws[f'B{row}'].font = styles['data_font']
        ws[f'A{row}'].alignment = styles['data_alignment']
        ws[f'B{row}'].alignment = styles['currency_alignment']
        
        # Apply borders
        for row_num in range(1, row + 1):
            for col in ['A', 'B']:
                cell = ws[f'{col}{row_num}']
                cell.border = styles['border']
        
        # Auto-fit columns
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15
        
        # Freeze panes
        ws.freeze_panes = 'A5'
        
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    @staticmethod
    def export_cash_flow_pdf(data: Dict[str, Any], org_name: str, start_date: date, end_date: date) -> BytesIO:
        """Export Cash Flow Statement to PDF"""
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=1*inch)
        story = []
        
        title_style, subtitle_style, section_style = ReportExportService._get_pdf_styles()
        
        # Title and subtitle
        story.append(Paragraph("CASH FLOW STATEMENT", title_style))
        story.append(Paragraph(f"{org_name}", subtitle_style))
        story.append(Paragraph(f"For the period {start_date.strftime('%B %d, %Y')} to {end_date.strftime('%B %d, %Y')}", subtitle_style))
        story.append(Spacer(1, 20))
        
        # Operating Activities section
        story.append(Paragraph("OPERATING ACTIVITIES", section_style))
        operating_data = [["Description", "Amount"]]
        operating_data.append(["Net Cash from Operating Activities", ReportExportService._format_currency(data.get("operating_cash_flow", "0"))])
        operating_table = ReportExportService._create_pdf_table(operating_data)
        story.append(operating_table)
        story.append(Spacer(1, 12))
        
        # Investing Activities section
        story.append(Paragraph("INVESTING ACTIVITIES", section_style))
        investing_data = [["Description", "Amount"]]
        investing_data.append(["Net Cash from Investing Activities", ReportExportService._format_currency(data.get("investing_cash_flow", "0"))])
        investing_table = ReportExportService._create_pdf_table(investing_data)
        story.append(investing_table)
        story.append(Spacer(1, 12))
        
        # Financing Activities section
        story.append(Paragraph("FINANCING ACTIVITIES", section_style))
        financing_data = [["Description", "Amount"]]
        financing_data.append(["Net Cash from Financing Activities", ReportExportService._format_currency(data.get("financing_cash_flow", "0"))])
        financing_table = ReportExportService._create_pdf_table(financing_data)
        story.append(financing_table)
        story.append(Spacer(1, 12))
        
        # Net Change in Cash
        story.append(Paragraph("NET CHANGE IN CASH", section_style))
        net_change_data = [["Description", "Amount"]]
        net_change_data.append(["Net Change in Cash", ReportExportService._format_currency(data.get("net_cash_flow", "0"))])
        
        # Special styling for net change
        net_change_table = Table(net_change_data, colWidths=[3*inch, 1.5*inch])
        net_change_style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 1), (-1, 1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 1), (-1, 1), 12),
            ('ALIGN', (0, 1), (0, 1), 'LEFT'),
            ('ALIGN', (1, 1), (1, 1), 'RIGHT'),
            ('BACKGROUND', (0, 1), (-1, 1), colors.lightgreen),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ])
        net_change_table.setStyle(net_change_style)
        story.append(net_change_table)
        
        doc.build(story)
        buffer.seek(0)
        return buffer

    @staticmethod
    def export_cash_flow_excel(data: Dict[str, Any], org_name: str, start_date: date, end_date: date) -> BytesIO:
        """Export Cash Flow Statement to Excel"""
        buffer = BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = "Cash Flow"
        
        styles = ReportExportService._get_excel_styles()
        
        # Set page margins
        ws.page_margins = PageMargins(left=0.7, right=0.7, top=0.75, bottom=0.75)
        
        # Title
        ws['A1'] = "CASH FLOW STATEMENT"
        ws['A1'].font = Font(name='Arial', size=16, bold=True, color='366092')
        ws['A1'].alignment = Alignment(horizontal='center')
        ws.merge_cells('A1:B1')
        
        # Organization and period
        ws['A2'] = org_name
        ws['A2'].font = Font(name='Arial', size=12, bold=True)
        ws['A2'].alignment = Alignment(horizontal='center')
        ws.merge_cells('A2:B2')
        
        ws['A3'] = f"For the period {start_date.strftime('%B %d, %Y')} to {end_date.strftime('%B %d, %Y')}"
        ws['A3'].font = Font(name='Arial', size=10)
        ws['A3'].alignment = Alignment(horizontal='center')
        ws.merge_cells('A3:B3')
        
        # Operating Activities section
        row = 5
        ws[f'A{row}'] = "OPERATING ACTIVITIES"
        ws[f'A{row}'].font = Font(name='Arial', size=12, bold=True, color='366092')
        row += 1
        
        ws[f'A{row}'] = "Net Cash from Operating Activities"
        ws[f'B{row}'] = float(data.get("operating_cash_flow", "0"))
        ws[f'B{row}'].number_format = '"$"#,##0.00'
        ws[f'A{row}'].font = styles['data_font']
        ws[f'B{row}'].font = styles['currency_font']
        ws[f'A{row}'].alignment = styles['data_alignment']
        ws[f'B{row}'].alignment = styles['currency_alignment']
        row += 2
        
        # Investing Activities section
        ws[f'A{row}'] = "INVESTING ACTIVITIES"
        ws[f'A{row}'].font = Font(name='Arial', size=12, bold=True, color='366092')
        row += 1
        
        ws[f'A{row}'] = "Net Cash from Investing Activities"
        ws[f'B{row}'] = float(data.get("investing_cash_flow", "0"))
        ws[f'B{row}'].number_format = '"$"#,##0.00'
        ws[f'A{row}'].font = styles['data_font']
        ws[f'B{row}'].font = styles['currency_font']
        ws[f'A{row}'].alignment = styles['data_alignment']
        ws[f'B{row}'].alignment = styles['currency_alignment']
        row += 2
        
        # Financing Activities section
        ws[f'A{row}'] = "FINANCING ACTIVITIES"
        ws[f'A{row}'].font = Font(name='Arial', size=12, bold=True, color='366092')
        row += 1
        
        ws[f'A{row}'] = "Net Cash from Financing Activities"
        ws[f'B{row}'] = float(data.get("financing_cash_flow", "0"))
        ws[f'B{row}'].number_format = '"$"#,##0.00'
        ws[f'A{row}'].font = styles['data_font']
        ws[f'B{row}'].font = styles['currency_font']
        ws[f'A{row}'].alignment = styles['data_alignment']
        ws[f'B{row}'].alignment = styles['currency_alignment']
        row += 2
        
        # Net Change in Cash section
        ws[f'A{row}'] = "NET CHANGE IN CASH"
        ws[f'A{row}'].font = Font(name='Arial', size=12, bold=True, color='366092')
        row += 1
        
        ws[f'A{row}'] = "Net Change in Cash"
        ws[f'B{row}'] = float(data.get("net_cash_flow", "0"))
        ws[f'B{row}'].number_format = '"$"#,##0.00'
        ws[f'A{row}'].font = styles['total_font']
        ws[f'B{row}'].font = styles['total_font']
        ws[f'A{row}'].fill = styles['total_fill']
        ws[f'B{row}'].fill = styles['total_fill']
        ws[f'A{row}'].alignment = styles['data_alignment']
        ws[f'B{row}'].alignment = styles['currency_alignment']
        
        # Apply borders
        for row_num in range(1, row + 1):
            for col in ['A', 'B']:
                cell = ws[f'{col}{row_num}']
                cell.border = styles['border']
        
        # Auto-fit columns
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15
        
        # Freeze panes
        ws.freeze_panes = 'A5'
        
        wb.save(buffer)
        buffer.seek(0)
        return buffer
